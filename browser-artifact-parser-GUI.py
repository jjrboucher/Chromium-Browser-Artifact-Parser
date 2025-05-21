# Written by Jacques Boucher (Enhanced UI Version)
# email: jjrboucher@gmail.com
# version date: 2025-May-20 (UI Enhanced)
#
# Script to extract data from Google Chrome's or MS Edge's SQLite databases
# Outputs to an Excel file with modern UI and artifact selection
#
# tested with Chrome 129, Edge 129, Opera 113

# Written by Jacques Boucher (Enhanced UI Version)
# email: jjrboucher@gmail.com
# version date: 2025-Mar-05 (UI Enhanced)
#
# Script to extract data from Google Chrome's or MS Edge's SQLite databases
# Outputs to an Excel file with modern UI and artifact selection
#
# tested with Chrome 129, Edge 129, Opera 113

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, font
from Classes.Preferences import Preferences
from Functions.write_to_excel import write_excel
from JSON.bookmarks import get_chromium_bookmarks
from SQLite.cookies import chrome_cookies
from SQLite.downloads import chrome_downloads, chrome_downloads_gaps
from SQLite.favicons import chrome_favicons
from SQLite.history import chrome_history, chrome_history_gaps
from SQLite.logindata import chrome_login_data, chrome_login_data_gaps
from SQLite.searchterms import chrome_keyword_historyquery
from SQLite.shortcuts import chrome_shortcuts
from SQLite.topsites import chrome_topsites
from SQLite.WebData import (
    chrome_autofill,
    chrome_keywords,
    chrome_masked_credit_cards,
    chrome_masked_bank_accounts,
    chrome_addresses
)
from SQLite.webasssist import edge_webassist

import openpyxl
import pandas as pd
import sqlite3
import numpy as np
import io
import threading


class ModernChromeParserGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Chromium Browser Parser - Enhanced Edition")
        self.root.geometry("1000x750")
        self.root.configure(bg='#f0f0f0')

        # Style configuration
        self.setup_styles()

        # Initialize variables
        self.profile_path = None
        self.output_path = None
        self.is_processing = False

        # Artifact selection variables
        self.artifact_vars = {}
        self.setup_artifact_selection()

        # Create the main UI
        self.create_widgets()

        # Try to set icon (handle gracefully if not found)
        try:
            self.icon = tk.PhotoImage(file="./images/browser_chromium_icon.png")
            self.root.iconphoto(False, self.icon)
        except:
            pass  # Continue without icon if file not found

    def setup_styles(self):
        """Configure modern styling for the application"""
        self.style = ttk.Style()

        # Configure styles for different widgets
        self.style.configure('Title.TLabel',
                             font=('Segoe UI', 16, 'bold'),
                             background='#f0f0f0',
                             foreground='#2c3e50')

        self.style.configure('Section.TLabel',
                             font=('Segoe UI', 12, 'bold'),
                             background='#f0f0f0',
                             foreground='#34495e')

        self.style.configure('Modern.TButton',
                             padding=(6, 2),
                             font=('Segoe UI', 10))

        self.style.configure('Action.TButton',
                             padding=(8, 4),
                             font=('Segoe UI', 11, 'bold'))

        # Progress bar styling - simplified for compatibility
        try:
            self.style.configure('Vertical.TProgressbar',
                                 thickness=20)
        except:
            pass  # Use default styling if custom styling fails

    def setup_artifact_selection(self):
        """Initialize artifact selection options"""
        self.artifacts_config = {
            'History': {'enabled': True, 'query': 'History'},
            'History Gaps': {'enabled': True, 'query': 'History Gaps'},
            'Downloads': {'enabled': True, 'query': 'Downloads'},
            'Downloads Gaps': {'enabled': True, 'query': 'Downloads Gaps'},
            'Autofill': {'enabled': True, 'query': 'Autofill'},
            'Addresses': {'enabled': True, 'query': 'Addresses'},
            'Keywords': {'enabled': True, 'query': 'Keywords'},
            'Credit Cards': {'enabled': True, 'query': 'Credit Cards'},
            'Bank Accounts': {'enabled': True, 'query': 'Bank Accounts'},
            'Login Data': {'enabled': True, 'query': 'Login Data'},
            'Login Data Gaps': {'enabled': True, 'query': 'Login Data Gaps'},
            'Shortcuts': {'enabled': True, 'query': 'Shortcuts'},
            'Top Sites': {'enabled': True, 'query': 'Top Sites'},
            'Cookies': {'enabled': True, 'query': 'Cookies'},
            'FavIcons': {'enabled': True, 'query': 'FavIcons'},
            'Search Terms': {'enabled': True, 'query': 'Search Terms'},
            'Bookmarks': {'enabled': True, 'query': 'Bookmarks'},
            'Preferences': {'enabled': True, 'query': 'Preferences'},
            'Web Assist (Edge)': {'enabled': False, 'query': 'Web Assist'}
        }

        # Create BooleanVar for each artifact
        for artifact in self.artifacts_config:
            self.artifact_vars[artifact] = tk.BooleanVar(
                value=self.artifacts_config[artifact]['enabled']
            )

    def create_widgets(self):
        """Create and arrange all GUI widgets"""
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights for responsive design
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(10, weight=1)  # Make status section expandable

        # Title
        title_label = ttk.Label(main_frame, text="🌐 Chromium Browser Parser",
                                style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # File paths section (rows 1-2)
        self.create_file_paths_section(main_frame, 1)

        # Artifact selection section (row 3-5)
        self.create_artifact_selection_section(main_frame, 3)

        # Progress and action section (row 6-7)
        self.create_progress_section(main_frame, 6)

        # Status section (row 8-9)
        self.create_status_section(main_frame, 8)

    def create_file_paths_section(self, parent, row):
        """Create the file paths input section"""
        # Profile path
        ttk.Label(parent, text="Chrome User Profile Folder:",
                  style='Section.TLabel').grid(row=row, column=0, sticky="w", pady=(0, 10))

        profile_frame = ttk.Frame(parent)
        profile_frame.grid(row=row, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        profile_frame.columnconfigure(0, weight=1)

        self.profile_entry = ttk.Entry(profile_frame, font=('Segoe UI', 10))
        self.profile_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))

        ttk.Button(profile_frame, text="📁 Browse", command=self.browse_profile,
                   style='Modern.TButton').grid(row=0, column=1)

        # Output path
        ttk.Label(parent, text="Output Excel File:",
                  style='Section.TLabel').grid(row=row + 1, column=0, sticky="w", pady=(0, 5))

        output_frame = ttk.Frame(parent)
        output_frame.grid(row=row + 1, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        output_frame.columnconfigure(0, weight=1)

        self.output_entry = ttk.Entry(output_frame, font=('Segoe UI', 10))
        self.output_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))

        ttk.Button(output_frame, text="💾 Save As", command=self.browse_output,
                   style='Modern.TButton').grid(row=0, column=1)

    def create_artifact_selection_section(self, parent, row):
        """Create the artifact selection checkboxes"""
        # Section title with appropriate spacing
        ttk.Label(parent, text="Select Artifacts to Process:",
                  style='Section.TLabel').grid(row=row, column=0, columnspan=3, sticky="w", pady=(0, 5))

        # Create scrollable frame for artifacts with fixed height
        artifact_frame = ttk.Frame(parent)
        artifact_frame.grid(row=row + 1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        artifact_frame.columnconfigure(0, weight=1)

        canvas = tk.Canvas(artifact_frame, height=100, bg='white', highlightthickness=1,
                           highlightbackground='#bdc3c7')
        scrollbar = ttk.Scrollbar(artifact_frame, orient="vertical", command=canvas.yview)


        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky=("nsew")) # tk.W, tk.E
        scrollbar.grid(row=0, column=1, sticky=("ns")) # tk.N, tk.S

        # Add checkboxes in a grid layout (5 columns)
        artifacts = list(self.artifacts_config.keys())
        for i, artifact in enumerate(artifacts):
            row_pos = i // 5
            col_pos = i % 5

            cb = ttk.Checkbutton(scrollable_frame, text=artifact,
                                 variable=self.artifact_vars[artifact],
                                 padding=(5, 2))
            cb.grid(row=row_pos, column=col_pos, sticky="w", padx=10, pady=0)

        # Select/Deselect all buttons
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=row + 2, column=0, columnspan=3, pady=(0, 0))

        ttk.Button(button_frame, text="✓ Select All",
                   command=self.select_all_artifacts,
                   style='Modern.TButton').grid(row=0, column=0, padx=(0, 10))

        ttk.Button(button_frame, text="✗ Deselect All",
                   command=self.deselect_all_artifacts,
                   style='Modern.TButton').grid(row=0, column=1)

    def create_progress_section(self, parent, row):
        """Create the progress bar and action buttons section"""
        # Add a separator line
        separator = ttk.Separator(parent, orient='horizontal')
        separator.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 5))

        progress_frame = ttk.Frame(parent)
        progress_frame.grid(row=row + 1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 10))
        progress_frame.columnconfigure(1, weight=1)

        # Vertical progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame,
                                            orient='horizontal',
                                            mode='determinate',
                                            variable=self.progress_var,
                                            style='Horizontal.TProgressbar')
        self.progress_bar.grid(row=0, column=2, columnspan=3, sticky=(tk.N, tk.S),
                               padx=(0, 20), pady=2)
        self.progress_bar.configure(length=400)

        # Progress label
        self.progress_label = ttk.Label(progress_frame, text="Ready to process",
                                        font=('Segoe UI', 10, 'italic'))
        self.progress_label.grid(row=0, column=1, sticky="w", pady=(0, 5))

        # Action buttons
        button_container = ttk.Frame(progress_frame)
        button_container.grid(row=0, column=1, sticky="w")

        self.run_button = ttk.Button(button_container, text="🚀 Run Parser",
                                     command=self.run_parser_threaded,
                                     style='Action.TButton')
        self.run_button.grid(row=0, column=0, padx=(0, 10))

        self.stop_button = ttk.Button(button_container, text="⏹ Stop",
                                      command=self.stop_processing,
                                      style='Action.TButton',
                                      state='disabled')
        self.stop_button.grid(row=0, column=1, padx=(0, 10))

        ttk.Button(button_container, text="❌ Exit",
                   command=self.root.destroy,
                   style='Action.TButton').grid(row=0, column=2)

    def create_status_section(self, parent, row):
        """Create the status display section"""

        ttk.Label(parent, text="Processing Status:",
                  style='Section.TLabel').grid(row=row + 1, column=0, columnspan=3, sticky="w", pady=(10, 5))

        status_frame = ttk.Frame(parent)
        status_frame.grid(row=row + 2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        status_frame.columnconfigure(0, weight=1)
        status_frame.rowconfigure(0, weight=1)

        # Status text with scrollbar
        self.status_text = tk.Text(status_frame, height=12, width=80,
                                   font=('Consolas', 9),
                                   bg='#2c3e50', fg='#ecf0f1',
                                   selectbackground='#34495e',
                                   wrap=tk.WORD,
                                   state="disabled")

        status_scrollbar = ttk.Scrollbar(status_frame, orient="vertical",
                                         command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)

        self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        status_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        # Initial welcome message
        self.update_status("Welcome to Chromium Browser Parser!")
        self.update_status("Select your Chrome profile folder and output file, then click 'Run Parser'.")

    def select_all_artifacts(self):
        """Select all artifact checkboxes"""
        for var in self.artifact_vars.values():
            var.set(True)

    def deselect_all_artifacts(self):
        """Deselect all artifact checkboxes"""
        for var in self.artifact_vars.values():
            var.set(False)

    def update_status(self, message):
        """Update the status text display"""
        self.status_text.config(state="normal")
        self.status_text.insert(tk.END, f"[{self.get_timestamp()}] {message}\n")
        self.status_text.see(tk.END)
        self.status_text.config(state="disabled")
        self.root.update_idletasks()

    def get_timestamp(self):
        """Get current timestamp for status messages"""
        import datetime
        return datetime.datetime.now().strftime("%H:%M:%S")

    def update_progress(self, current, total, message=""):
        """Update the progress bar and label"""
        if total > 0:
            percentage = (current / total) * 100
            self.progress_var.set(percentage)
            if message:
                self.progress_label.config(text=f"{message} ({current}/{total})")
            else:
                self.progress_label.config(text=f"Processing... {current}/{total} ({percentage:.1f}%)")
        else:
            self.progress_var.set(0)
            self.progress_label.config(text=message or "Ready to process")
        self.root.update_idletasks()

    def browse_profile(self):
        """Open file dialog to select Chrome profile folder"""
        self.profile_path = filedialog.askdirectory(
            title="Select Chrome User Profile Folder",
            mustexist=True
        )
        if self.profile_path:
            self.profile_entry.delete(0, tk.END)
            self.profile_entry.insert(0, self.profile_path)
            self.update_status(f"Selected profile: {self.profile_path}")

    def browse_output(self):
        """Open file dialog to select output Excel file"""
        self.output_path = filedialog.asksaveasfilename(
            title="Select Output Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if self.output_path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, self.output_path)
            self.update_status(f"Output file: {self.output_path}")

    def run_parser_threaded(self):
        """Run the parser in a separate thread to prevent UI freezing"""
        if self.is_processing:
            return

        # Validate inputs
        self.profile_path = self.profile_entry.get().strip()
        self.output_path = self.output_entry.get().strip()

        if not self.profile_path or not self.output_path:
            messagebox.showerror("Error", "Please select both profile and output paths.")
            return

        # Check if any artifacts are selected
        selected_artifacts = [name for name, var in self.artifact_vars.items() if var.get()]
        if not selected_artifacts:
            messagebox.showerror("Error", "Please select at least one artifact to process.")
            return

        # Start processing in separate thread
        self.is_processing = True
        self.run_button.config(state='disabled')
        self.stop_button.config(state='normal')

        self.processing_thread = threading.Thread(target=self.run_parser)
        self.processing_thread.daemon = True
        self.processing_thread.start()

    def stop_processing(self):
        """Stop the current processing (note: this is a basic implementation)"""
        self.is_processing = False
        self.update_status("Processing stopped by user.")
        self.run_button.config(state='normal')
        self.stop_button.config(state='disabled')
        self.update_progress(0, 1, "Stopped")

    def run_parser(self):
        """Main parser logic with progress tracking"""
        try:
            # Get selected artifacts
            selected_artifacts = [name for name, var in self.artifact_vars.items() if var.get()]
            total_artifacts = len(selected_artifacts)
            current_artifact = 0

            self.update_status(f"Starting to process {total_artifacts} selected artifacts...")
            self.update_progress(0, total_artifacts, "Initializing...")

            # Define query mappings
            chromium_queries = {
                'History': [f'{self.profile_path}/History', chrome_history],
                "History Gaps": [f'{self.profile_path}/History', chrome_history_gaps],
                "Downloads": [f'{self.profile_path}/History', chrome_downloads],
                "Downloads Gaps": [f'{self.profile_path}/History', chrome_downloads_gaps],
                "Autofill": [f'{self.profile_path}/Web Data', chrome_autofill],
                "Addresses": [f'{self.profile_path}/Web Data', chrome_addresses],
                "Keywords": [f'{self.profile_path}/Web Data', chrome_keywords],
                "Credit Cards": [f'{self.profile_path}/Web Data', chrome_masked_credit_cards],
                "Bank Accounts": [f'{self.profile_path}/Web Data', chrome_masked_bank_accounts],
                "Login Data": [f'{self.profile_path}/Login Data', chrome_login_data],
                "Login Data Gaps": [f'{self.profile_path}/Login Data', chrome_login_data_gaps],
                "Shortcuts": [f'{self.profile_path}/Shortcuts', chrome_shortcuts],
                "Top Sites": [f'{self.profile_path}/Top Sites', chrome_topsites],
                "Cookies": [f'{self.profile_path}/Network/Cookies', chrome_cookies],
                "FavIcons": [f'{self.profile_path}/Favicons', chrome_favicons]
            }

            edge_queries = {
                "Web Assist (Edge)": [f'{self.profile_path}/WebAssistDatabase', edge_webassist]
            }

            record_counts = []

            # Process SQLite-based artifacts
            for artifact_name in selected_artifacts:
                if not self.is_processing:
                    break

                current_artifact += 1
                self.update_progress(current_artifact, total_artifacts, f"Processing {artifact_name}")
                self.update_status(f"Processing {artifact_name}...")

                try:
                    # Check if it's a chromium or edge query
                    if artifact_name in chromium_queries:
                        df, ws = self.get_dataframes(
                            chromium_queries[artifact_name][0],
                            chromium_queries[artifact_name][1]
                        )
                        write_excel(df, ws, self.output_path)
                        record_counts.append((ws, len(df)))
                        self.update_status(f"✓ {artifact_name}: {len(df)} records processed")

                    elif artifact_name in edge_queries:
                        df, ws = self.get_dataframes(
                            edge_queries[artifact_name][0],
                            edge_queries[artifact_name][1]
                        )
                        write_excel(df, ws, self.output_path)
                        record_counts.append((ws, len(df)))
                        self.update_status(f"✓ {artifact_name}: {len(df)} records processed")

                    # Handle special cases
                    elif artifact_name == "Search Terms":
                        dataframe_searchterms, ws = self.process_search_terms()
                        write_excel(dataframe_searchterms, ws, self.output_path)
                        record_counts.append((ws, len(dataframe_searchterms)))
                        self.update_status(f"✓ Search Terms: {len(dataframe_searchterms)} records processed")

                    elif artifact_name == "Bookmarks":
                        bookmarks_df, ws = self.process_bookmarks()
                        write_excel(bookmarks_df, ws, self.output_path)
                        record_counts.append((ws, len(bookmarks_df)))
                        self.update_status(f"✓ Bookmarks: {len(bookmarks_df)} records processed")

                    elif artifact_name == "Preferences":
                        temp_record_count = self.process_preferences()
                        record_counts.append(("Preferences", temp_record_count))
                        self.update_status(f"✓ Preferences processed")

                except Exception as error:
                    self.update_status(f"❌ Failed to process {artifact_name}")
                    record_counts.append((artifact_name, 0))
                    if "database is locked" in str(error):
                        self.update_status(f"   Database file is locked. Close the browser and try again.")

            if self.is_processing:
                # Create summary
                self.update_status("Creating summary worksheet...")
                summary_df = pd.DataFrame(record_counts, columns=["Worksheet Name", "Record Count"])
                write_excel(summary_df, "Summary", self.output_path)

                # Reorganize workbook
                self.update_status("Organizing worksheets...")
                self.reorganize_workbook()

                self.update_status("✅ All processing completed successfully!")
                self.update_status(f"📁 Output saved to: {self.output_path}")
                self.update_progress(total_artifacts, total_artifacts, "Completed!")

                # Show completion message
                messagebox.showinfo("Success",
                                    f"Processing completed successfully!\n\nOutput saved to:\n{self.output_path}")
            else:
                self.update_status("⚠ Processing was stopped by user.")

        except Exception as e:
            self.update_status(f"❌ Critical error: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

        finally:
            # Reset UI state
            self.is_processing = False
            self.run_button.config(state='normal')
            self.stop_button.config(state='disabled')
            if not hasattr(self, 'progress_var') or self.progress_var.get() < 100:
                self.update_progress(0, 1, "Ready to process")

    def get_dataframes(self, db_file, function):
        """Get dataframes from SQLite database"""
        query, worksheet_name = function()
        conn = sqlite3.connect(db_file)
        dataframe = pd.read_sql_query(query, conn)
        conn.close()
        return dataframe, worksheet_name

    def process_search_terms(self):
        """Process search terms data"""
        worksheet = 'Search Terms'
        input_file = f'{self.profile_path}/History'
        df_history, ws_history = self.get_dataframes(input_file, chrome_keyword_historyquery)

        input_file = f'{self.profile_path}/Web Data'
        df_keywords, ws_keyword = self.get_dataframes(input_file, chrome_keywords)

        searchterms = []
        if len(df_keywords) > 0:
            for row in df_history.itertuples():
                if not np.isnan(row[3]):
                    try:
                        kw = [
                            df_keywords.query(f'id == {row[3]}')['keyword'].values[0],
                            df_keywords.query(f'id == {row[3]}')['date_created'].values[0],
                            df_keywords.query(f'id == {row[3]}')['Decoded date_created (UTC)'].values[0],
                            df_keywords.query(f'id == {row[3]}')['last_modified'].values[0],
                            df_keywords.query(f'id == {row[3]}')['Decoded last_modified (UTC)'].values[0]
                        ]
                    except IndexError:
                        kw = ['', '', '', '', '']
                    searchterms.append([
                        row[1], row[2], row[3], kw[0], row[5], row[6], kw[1], kw[2], kw[3], kw[4], row[7], row[8]
                    ])
        else:
            searchterms = [["", "", "", "", "", "", "", "", "", "", "", ""]]

        df_searchterms = pd.DataFrame(searchterms)
        df_searchterms.columns = [
            'URL id', 'url', 'keyword id', 'keyword', 'search term', 'typed_count', 'date_created',
            'Decoded date_created (UTC)', 'last_modified', 'Decoded last_modified (UTC)',
            'last_visit_time (UTC)', 'Decoded last_visit_time (UTC)'
        ]
        return df_searchterms, worksheet

    def process_bookmarks(self):
        """Process bookmarks data"""
        ws = 'Bookmarks'
        try:
            bookmarks_df, ws = get_chromium_bookmarks(f'{self.profile_path}/Bookmarks')
        except:
            bookmarks_df = pd.DataFrame()

        try:
            bookmarks_backup_df, ws_bak = get_chromium_bookmarks(f'{self.profile_path}/Bookmarks.bak')
        except:
            bookmarks_backup_df = pd.DataFrame()

        all_bookmarks = pd.concat([bookmarks_df, bookmarks_backup_df], ignore_index=True)
        return all_bookmarks, ws

    def process_preferences(self):
        """Process preferences data"""
        preferences = Preferences(f'{self.profile_path}/Preferences')
        preferences_output = io.StringIO()
        print(preferences, file=preferences_output)
        preferences_data = preferences_output.getvalue().splitlines()
        preferences_df = pd.DataFrame(preferences_data, columns=["Preferences Output"])
        write_excel(preferences_df, "Preferences", self.output_path)
        return len(preferences_df)


    def reorganize_workbook(self):
        """Reorganize the Excel workbook sheets"""
        wb = openpyxl.load_workbook(self.output_path)

        # Move important sheets to the front
        if 'Summary' in wb.sheetnames:
            current_index = wb.sheetnames.index(wb["Summary"].title)
            wb.move_sheet(wb["Summary"], -current_index)

        if 'Preferences' in wb.sheetnames:
            current_index = wb.sheetnames.index(wb["Preferences"].title)
            wb.move_sheet(wb["Preferences"], -current_index + 1)

        wb.save(self.output_path)


if __name__ == '__main__':
    # Create and configure the main window
    root = tk.Tk()
    app = ModernChromeParserGUI(root)

    # Make the window resizable
    root.resizable(True, True)

    # Set minimum window size
    root.minsize(700, 600)

    # Center the window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    # Start the application
    root.mainloop()